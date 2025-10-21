from datetime import datetime

import openpyxl
from youtube_statisrics import YTstats
from openpyxl import Workbook
from config import get_youtube_api_key

# 비디오ID를 입력 -> 메타데이터를 엑셀파일로 저장

wb = openpyxl.Workbook()
wb.create_sheet("Test")
sheet = wb["Test"]
wb_sheet = wb["Sheet"]

# wbid = openpyxl.load_workbook('/Users/YUHWAN PARK/Desktop/youtube_matadata/videolist.xlsx')
# wbid_sheet = wbid['Sheet1']

# print(wbid_sheet['A{}'.format(1)].value)

API_KEY = get_youtube_api_key()
videolist =["S_0me7vYyeU"]
wb_sheet.cell(1,1,datetime.now())
now = datetime.now()
i=1
datas = videolist
sheet.cell( i,1, 'view') # row, column, value
sheet.cell( i,2, 'like') # row, column, value
sheet.cell( i,3, 'comment') # row, column, value
sheet.cell( i,4, 'length') # row, column, value
sheet.cell( i,5, 'caption') # row, column, value
sheet.cell( i,6, 'postingperiod') # row, column, value
sheet.cell( i,7, 'subcount') # row, column, value
sheet.cell( i,8, 'video_id') # row, column, value
sheet.cell( i,9, 'channel_id') # row, column, value
sheet.cell( i,10, 'tags') # row, column, value
sheet.cell( i,11, 'categoryId') # row, column, value
sheet.cell( i,12, 'gold') # row, column, value
sheet.cell( i,13, 'silver')
sheet.cell( i,14, 'shortvideo')
sheet.cell( i,15, 'channel_title')
sheet.cell( i,16, 'video_title')
sheet.cell( i,17, 'postingdate')
sheet.cell( i,18, 'postingday')
print('get video data...')
for a in datas:

    # video_id = (wbid_sheet['A{}'.format(a)].value)
    video_id = a


    yt = YTstats(API_KEY, video_id)

    data = yt._get_single_video_data(video_id,part='snippet')#채널 id 추출, 개재 시간 추출
    data = yt._get_single_video_data(video_id,part='statistics')# 조회수, 좋아요수, 댓글수 추출
    data = yt._get_single_video_data(video_id,part='contentDetails') # 영상길이, 자막 
    yt.dump()
    data = yt._get_single_video_data(video_id,part='snippet')
    cid = data['channelId']
    postingdate =data['publishedAt']
    description = data['description']
    tags = description.count('#')
    categoryId = data['categoryId']
    channelTitle = data['channelTitle']
    title=data['title']

    data = yt._get_single_video_data(video_id,part='statistics')
    try:
        like = data['likeCount']
    except:
        like = 'null'
    view = data['viewCount']
    try:
        comment = data['commentCount']
    except:
        comment = 0
    data = yt._get_single_video_data(video_id,part='contentDetails')
    length =  data['duration']
    length = length[2:]
    length = length.replace("H","*3600+")
    length = length.replace("S","")
    length = length.replace("M","*60+")
    try:
        length = eval(length)
    except:
        try:
            length = eval(length[:-1])
        except:
            continue


    caption = data['caption']
    #============================= 채널 정보 구하기
    channel_id = cid
    yt = YTstats(API_KEY, channel_id)

    data = yt.get_channel_statistics()

    try:
        subcount = data['subscriberCount']
    except:
        subcount = 'null'
    print("채널 아이디 : ",cid)
    print("업로드 날짜 : ",postingdate)
    print("좋아요 수 : ",like)
    print("조회 수 : ",view)
    print("댓글 수 : ",comment)
    print("영상 길이: ",length)
    print("자막의 유무 : ",caption)
    print("구독자 수 :", subcount)
    print("영상 아이디", video_id)
    print(i,"번")
    #==============================업로드 기간 구하기 
    past = datetime.strptime(postingdate[0:10],"%Y-%m-%d")
    now = datetime.now()
    diff = now - past
    postingperiod = str(diff.days) + "일"
    date = postingdate[0:10]
    day = past.strftime("%A")

    print(postingperiod)


    sheet.cell( i+1,1, view) # row, column, value
    sheet.cell( i+1,2, like) # row, column, value
    sheet.cell( i+1,3, comment) # row, column, value
    sheet.cell( i+1,4, length) # row, column, value
    sheet.cell( i+1,5, caption) # row, column, value
    sheet.cell( i+1,6, postingperiod) # row, column, value
    sheet.cell( i+1,7, subcount) # row, column, value
    sheet.cell( i+1,8, video_id) # row, column, value
    sheet.cell( i+1,9, cid) # row, column, value
    sheet.cell( i+1,10, tags) # row, column, value
    sheet.cell( i+1,11, categoryId) # row, column, value
    #===============================================골드 버튼 / 실버버튼
    if subcount == 'null':
        sheet.cell( i+1,12, 0) # row, column, value
        sheet.cell( i+1,13, 0) # row, column, value
    elif int(subcount) >= 1000000 :
        sheet.cell( i+1,12, 1) # row, column, value
        sheet.cell( i+1,13, 0) # row, column, value
    elif 100000 <= int(subcount) <= 1000000 :
        sheet.cell( i+1,12, 0) # row, column, value
        sheet.cell( i+1,13, 1) # row, column, value
    else :
        sheet.cell( i+1,12, 0) # row, column, value
        sheet.cell( i+1,13, 0) # row, column, value
    if length <= 60:
        sheet.cell( i+1,14, 1)
    else:
        sheet.cell( i+1,14, 0)
    sheet.cell( i+1,15,channelTitle) # row, column, value
    sheet.cell( i+1,16,title) # row, column, value
    sheet.cell( i+1,17,date) # row, column, value
    sheet.cell( i+1,18,day) # row, column, value
    print(i,end=" ")

    i += 1
    
print("")
wb_sheet.cell(1,2,datetime.now())
wb.save('video_data.xlsx')
print(i-1,"개의 데이터 수집 완료")
print("end")